# -*- coding: utf-8 -*-
"""
Created on Tue Jul 21 16:36:24 2020

@author: kwangil_kim1
"""



from PIL import Image
import os, glob, sys, numpy as np
from sklearn.model_selection import train_test_split
import tensorflow as tf

from tensorflow.keras.models import Sequential
from tensorflow.keras.layers import Conv2D, MaxPooling2D, Dense, Flatten, Dropout
from tensorflow.keras.callbacks import EarlyStopping, ModelCheckpoint
import matplotlib.pyplot as plt

from PIL import Image
import os, glob, numpy as np
from tensorflow.keras.models import load_model

import cv2
import shutil
from openpyxl import load_workbook
from openpyxl import Workbook
import pandas as pd
from datetime import datetime


# 결점 name(공백없는 영문표기)
input_set_defect_pre = "R10"
input_set_defect = "R13"


# img_dir : 입력 이미지 상위 경로 
# img_dir = 'D:/CASTING/' + input_set_defect + '/analysis_split_target_image'
img_dir_pre = 'D:/CASTING/' + input_set_defect_pre + '/analysis_split_target_image_result/result/no_defect'
img_dir = 'D:/CASTING/' + input_set_defect + '/analysis_split_target_image'
# img_dir = 'D:/CPI_image_raw/B0605C/B0605C_image/20200605_214415_test'
# img_dir = 'D:/CPI_image_raw/B0608F/B0608F_Image/20200608_225258_test'
# img_dir = 'D:/CPI_image_raw/B0609A/B0609A_image/20200609_101130_test'
# img_dir = 'D:/CPI_image_raw/B0609B/B0609B_image/20200609_213003_test'
# img_dir = 'D:/CPI_image_raw/B0610A/B0610A_image/20200610_084852_test'
# img_dir = 'D:/CPI_image_raw/B0610C/B0610B_image/20200610_203848_test'


# 모델 버전 
test_ver = '08a'

# 모델 경로 설정 
img_dir_model = 'D:/CPI_model/CASTING_model/' + input_set_defect
categories = [input_set_defect, 'no_' + input_set_defect]
np_classes = len(categories)

# 이미지 사이즈
# image_w = 256
# image_h = 256
image_w = 64
image_h = 64


############################################## filename에 엑셀 파일 이름 작성 하면 됩니다.###################
filename = 'D:/CASTING/' + input_set_defect_pre + '/analysis_split_target_image/검사기_결과.csv'
#############################################################################################################


size_str = str(image_w)

model_name = 'CASTING_v1'

# model_dir = img_dir_model + '/model' + test_ver

model_path = img_dir_model +'/' + model_name

model = load_model(model_path +'/' + model_name + '.h5')









seed = 5
np.random.seed(seed)

def mkdir(path):
    if os.path.exists(path) is False:
        os.mkdir(path)


###
# 다음 추가학습을 위한 경로 지정 필요  
###


RESULT_SAVE_PATH_00 = img_dir + '_result'


RESULT_SAVE_PATH_10 = RESULT_SAVE_PATH_00 +'/result'
RESULT_SAVE_PATH_21 = RESULT_SAVE_PATH_10 + '/defect'
RESULT_SAVE_PATH_22 = RESULT_SAVE_PATH_10 + '/no_defect'
RESULT_SAVE_PATH_23 = RESULT_SAVE_PATH_10 + '/check'

mkdir(RESULT_SAVE_PATH_00)
mkdir(RESULT_SAVE_PATH_10)
mkdir(RESULT_SAVE_PATH_21)
mkdir(RESULT_SAVE_PATH_22)
mkdir(RESULT_SAVE_PATH_23)

####################################################################################################
# 성능 검사용 코드 
print('이미지 파일들을 전처리중 입니다.')
X = []
filenames = []
files = glob.glob(img_dir_pre + "/*.BMP")
for i, f in enumerate(files):
    img = Image.open(f)
    img = img.convert("RGB")
    img = img.resize((image_w, image_h))
    data = np.asarray(img)
    filenames.append(f)
    X.append(data)

X = np.array(X)

X = X.astype(float) / 255

# 엑셀파일에 컬럼 추가 

df = pd.read_csv(filename, encoding='CP949',index_col = '결함번호')
# df = pd.read_csv(filename,index_col = '결함번호')
df['분석 결과'] = ""

datetime_str = datetime.today().strftime("%Y%m%d%H%M%S")

# wb = load_workbook(filename)
# ws = wb.active
# ws.insert_cols(1) #첫번쨰 열(A열) 삽입

# 이미지 판별 및 내용 로그창에 출력 
print('전처리된 이미지 파일을 분류중 입니다.')
prediction = model.predict(X)
np.set_printoptions(formatter={'float': lambda x: "{0:0.3f}".format(x)})
cnt = 0
for i in prediction:
    # print(i, cnt, filenames[cnt].split("\\")[1])
    try:
        filename_index = int(filenames[cnt].split("\\")[1].split('PC1-')[1].split('.BMP')[0])
    except:
        filename_index = int(filenames[cnt].split("\\")[1].split('PC1-')[0].split('.BMP')[0])
        
    try :
        if i >= 0.7: 
            # if df['결함랭크(1～10)'][cnt+1] != 'S1' and df['결함랭크(1～10)'][cnt+1] != 'S2':
            if df['결함랭크(1～10)'][filename_index] != 'S1' and df['결함랭크(1～10)'][filename_index] != 'S2':
                print("해당 defect " + filenames[cnt].split("\\")[1] + "  이미지는 no_defect 로 추정됩니다. %s" % i)
                shutil.copy2(files[cnt], RESULT_SAVE_PATH_22 + '/' + filenames[cnt].split("\\")[1])
                # df['분석 결과'][cnt+1] = ''
                df['분석 결과'][filename_index] = ''
        elif i > 0.3: 
            # if df['결함랭크(1～10)'][cnt+1] != 'S1' and df['결함랭크(1～10)'][cnt+1] != 'S2':
            if df['결함랭크(1～10)'][filename_index] != 'S1' and df['결함랭크(1～10)'][filename_index] != 'S2':
                print("해당 defect " + filenames[cnt].split("\\")[1] + "  이미지는 check 으로 추정됩니다. %s" % i)
                shutil.copy2(files[cnt], RESULT_SAVE_PATH_23 + '/' + filenames[cnt].split("\\")[1])
                #ws.cell(row=cnt+2, column=1).value = 'check'
                # df['분석 결과'][cnt+1] = 'check'
                df['분석 결과'][filename_index] = 'check'
        else : 
            # if df['결함랭크(1～10)'][cnt+1] != 'S1' and df['결함랭크(1～10)'][cnt+1] != 'S2':
            if df['결함랭크(1～10)'][filename_index] != 'S1' and df['결함랭크(1～10)'][filename_index] != 'S2':
                print("해당 defect " + filenames[cnt].split("\\")[1] + "  이미지는 defect 으로 추정됩니다. %s" % i)
                shutil.copy2(files[cnt], RESULT_SAVE_PATH_21 + '/' + filenames[cnt].split("\\")[1])
                #ws.cell(row=cnt+2, column=1).value = 'defect'
                # df['분석 결과'][cnt+1] = 'defect'
                df['분석 결과'][filename_index] = 'defect'
        # os.remove(files[cnt])
    except:
        print(str(cnt+1) +"파일 없음 ")
    cnt += 1


#wb.save(filename)
dummy = pd.DataFrame()
dummy = df[['분석 결과','결함랭크(1～10)']]
dummy.to_csv(RESULT_SAVE_PATH_10 + '/검사기_결과_1' + '_' + input_set_defect + '_' + datetime_str + '.csv')

